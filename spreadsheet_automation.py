import openpyxl as xl
from openpyxl.chart import BarChart,Reference
def Process_workbook(filename):  # filename -> excet sheet on which you want to work 
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2,sheet.max_row + 1):
        cell = sheet.cell(row,3)
        corrected_price = cell.value*0.9
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price
        
    Max_row = sheet.max_row
    Max_col = sheet.max.col
    Min_row = sheet.mon_row
    Min_col = sheet.min_col
    values = Reference(sheet, min_row=Min_row, max_row=Max_row, min_col=Min_col, max_col=Max_col)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'e2')
    wb.save(filename)

    
Process_workbook("")
